#!/usr/bin/env python3
"""
CloudUnify Pro - One-off dataset ingestion utility

This script ingests attached Excel datasets (AWS, Azure, GCP resources, and recommendations)
into the Postgres database used by CloudUnify Pro.

What it does
- Discovers and parses all .xlsx files in the provided input directory (default: repo_root/attachments)
- Maps and transforms input columns to match target DB schemas for providers (aws/azure/gcp)
- Ensures/creates lookup rows as needed:
  - organizations
  - cloud_accounts
- Performs idempotent upserts:
  - resources keyed by (organization_id, provider, resource_id)
  - recommendations keyed by 'id' (uses recommendation_id, or stable UUIDv5 fallback when absent)
  - resource_costs_daily (ONLY if the table exists) keyed by 'id' using stable UUIDv5
- Handles UUID generation where absent (resources PK, auto-created lookups, missing recommendation IDs)
- Preserves referential integrity:
  - resources.cloud_account_id must exist (ensured/created)
  - recommendations.resource_id is resolved to internal resources.id when possible
  - resource_costs_daily.resource_id is the internal resources.id (from the upsert RETURNING)

DB connection precedence (per task requirement)
1) --database-url CLI argument
2) db_connection.txt (if present)
3) DATABASE_URL environment variable
4) fallback placeholder DSN constant (will fail unless replaced)

Neon notes
- The script enforces `sslmode=require` and `channel_binding=require` in the DSN.

Run locally
    cd cloudunify-pro-v1-287143-293501/BackendAPI
    python scripts/ingest_datasets.py --org <ORG_UUID> --cloud-account-id <ACCOUNT_UUID>

Tip: Use the Makefile target added in this change:
    make ingest-excels ORG_ID=<ORG_UUID> CLOUD_ACCOUNT_ID=<ACCOUNT_UUID>

"""

from __future__ import annotations

import argparse
import logging
import math
import os
import re
import sys
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

# Dependencies: psycopg (v3), openpyxl, python-dotenv (already in project requirements)
try:
    import psycopg
    from psycopg.rows import dict_row
    from psycopg.types.json import Json
except Exception:  # pragma: no cover
    print("ERROR: psycopg is required to run this script. Please install dependencies.", file=sys.stderr)
    raise

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    print("ERROR: openpyxl is required to parse Excel files. Please install dependencies.", file=sys.stderr)
    raise

# Pandas is optional: we will use it when available but do not require it.
try:  # pragma: no cover (optional dependency)
    import pandas as pd  # type: ignore
except Exception:  # pragma: no cover
    pd = None  # type: ignore

try:
    from dotenv import load_dotenv  # provided by project requirements
except Exception:

    def load_dotenv(*args: Any, **kwargs: Any) -> None:
        return


# ---------- Logging ----------

logger = logging.getLogger("cloudunify.scripts.ingest")
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s: %(message)s"))
logger.addHandler(handler)
logger.setLevel(logging.INFO)


# ---------- Constants and helpers ----------

# Replace this with the provided Neon connection string if neither CLI nor db_connection.txt nor env is provided.
# Do not commit secrets to version control in real deployments.
FALLBACK_NEON_DSN = "postgresql://USER:PASSWORD@HOST:PORT/DBNAME?sslmode=require"

PROVIDER_SLUGS = {"aws", "azure", "gcp"}
PRIORITY_ALLOWED = {"low", "medium", "high", "critical"}

# Stable UUID namespaces (constants) for deterministic ID generation
NAMESPACE_ORG = uuid.UUID("a9d8e4a4-9f16-4f6b-a8ef-8c0f8d54cf5a")
NAMESPACE_CLOUD_ACCOUNT = uuid.UUID("0cb9f6d1-2b6b-4e25-bc05-7c0a8d1c5fb1")
NAMESPACE_RECOMMENDATION = uuid.UUID("1d4ab8e8-0d5e-4c1a-9a41-2a30f1a34a2c")
NAMESPACE_RESOURCE_COST_DAILY = uuid.UUID("45dffbda-0f3f-4c9f-a9a9-3fbefb13e531")


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _is_missing(v: Any) -> bool:
    """Treat None, empty strings, NaN, and pandas NA-like values as missing."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    # pandas.NA / NaT support (without importing pandas types directly)
    try:
        if pd is not None and pd.isna(v):  # type: ignore[attr-defined]
            return True
    except Exception:
        pass
    return False


def _normalize_priority(v: Any) -> str:
    if _is_missing(v):
        return "medium"
    s = str(v).strip().lower()
    return s if s in PRIORITY_ALLOWED else "medium"


def _normalize_provider(v: Any) -> Optional[str]:
    if _is_missing(v):
        return None
    s = str(v).strip().lower()
    # Allow AWS/Azure/GCP variants
    if s in PROVIDER_SLUGS:
        return s
    if s in {"amazon", "aws"}:
        return "aws"
    if s in {"microsoft", "azure"}:
        return "azure"
    if s in {"google", "gcp"}:
        return "gcp"
    return None


def _provider_from_filename(name: str) -> Optional[str]:
    lower = name.lower()
    for prov in PROVIDER_SLUGS:
        if prov in lower:
            return prov
    return None


def _parse_tags(s: Any) -> Dict[str, str]:
    """
    Parse tag string like "Environment:production,Team:backend" into dict.
    Accepts dict (pass-through), None/NA (empty dict), list/tuple of "k:v", or string.
    """
    if _is_missing(s):
        return {}
    if isinstance(s, dict):
        return {str(k): str(v) for k, v in s.items() if not _is_missing(k)}
    if isinstance(s, (list, tuple)):
        out: Dict[str, str] = {}
        for item in s:
            if _is_missing(item):
                continue
            parts = str(item).split(":", 1)
            if len(parts) == 2:
                k, v = parts
                out[k.strip()] = v.strip()
        return out

    text = str(s).strip()
    if not text or text.lower() == "nan":
        return {}

    out: Dict[str, str] = {}
    for pair in text.split(","):
        if not pair.strip():
            continue
        if ":" in pair:
            k, v = pair.split(":", 1)
            out[k.strip()] = v.strip()
        else:
            out[pair.strip()] = "true"
    return out


def _to_decimal_or_none(v: Any) -> Optional[Decimal]:
    if _is_missing(v):
        return None
    if isinstance(v, Decimal):
        return v
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _parse_datetime_or_none(v: Any) -> Optional[datetime]:
    if _is_missing(v):
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)

    s = str(v).strip()
    if not s:
        return None
    # normalize trailing Z to UTC
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _snake(s: str) -> str:
    return s.strip().lower().replace(" ", "_")


def _normalize_state(state: Any) -> str:
    """Normalize provider state strings to a consistent, UI-friendly representation."""
    if _is_missing(state):
        return "unknown"
    return str(state).strip().lower()


def _slugify(s: str) -> str:
    """Simple slugify for organization slugs."""
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or "default"


def _ensure_ssl_and_channel_binding(dsn: str) -> str:
    """
    Ensure DSN has sslmode=require and channel_binding=require.
    Also normalize scheme for psycopg if SQLAlchemy async URLs are passed (e.g., postgresql+asyncpg://).
    """
    # Normalize scheme if using sqlalchemy URLs
    if dsn.startswith("postgresql+"):
        dsn = "postgresql://" + dsn.split("postgresql+", 1)[1]
    parts = urlparse(dsn)
    q = dict(parse_qsl(parts.query, keep_blank_values=True))
    if q.get("sslmode") is None:
        q["sslmode"] = "require"
    if q.get("channel_binding") is None:
        q["channel_binding"] = "require"
    new_parts = parts._replace(query=urlencode(q))
    return urlunparse(new_parts)


def _resolve_repo_root(start: Path) -> Optional[Path]:
    # Walk upward to find a directory that has 'attachments' as a child
    cur = start
    for _ in range(8):
        if (cur / "attachments").is_dir():
            return cur
        if cur.parent == cur:
            break
        cur = cur.parent
    return None


def _find_db_connection_txt(start: Path) -> Optional[Path]:
    """Find db_connection.txt by walking upward from a starting directory."""
    cur = start
    for _ in range(10):
        candidate = cur / "db_connection.txt"
        if candidate.exists() and candidate.is_file():
            return candidate
        if cur.parent == cur:
            break
        cur = cur.parent
    return None


def _parse_db_connection_txt(text: str) -> Optional[str]:
    """
    Parse db_connection.txt content.
    Supports formats like:
      psql postgresql://user:pass@host:5432/db?sslmode=require
    or a bare postgresql:// URL.
    """
    line = ""
    for raw in text.splitlines():
        raw = raw.strip()
        if not raw or raw.startswith("#"):
            continue
        line = raw
        break
    if not line:
        return None

    if line.startswith("psql "):
        parts = line.split()
        # typically: ["psql", "<dsn>"]
        if len(parts) >= 2:
            return parts[-1].strip().strip("'").strip('"')
    if line.startswith("postgresql://") or line.startswith("postgres://") or line.startswith("postgresql+"):
        return line.strip().strip("'").strip('"')
    return None


def _fetch_one_scalar(conn: psycopg.Connection, query: str, params: Tuple[Any, ...]) -> Optional[Any]:
    with conn.cursor() as cur:
        cur.execute(query, params)
        row = cur.fetchone()
        return row[0] if row and len(row) >= 1 else None


def _table_exists(conn: psycopg.Connection, table_name: str, schema: str = "public") -> bool:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema = %s AND table_name = %s
            LIMIT 1
            """,
            (schema, table_name),
        )
        return cur.fetchone() is not None


def _get_table_columns(conn: psycopg.Connection, table_name: str, schema: str = "public") -> List[str]:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = %s AND table_name = %s
            ORDER BY ordinal_position
            """,
            (schema, table_name),
        )
        return [r[0] for r in cur.fetchall()]


@dataclass
class TableReport:
    inserted: int = 0
    updated: int = 0
    skipped: int = 0


@dataclass(frozen=True)
class IngestConfig:
    organization_id: str
    account_by_provider: Dict[str, str]
    dsn: str
    input_dir: Path
    dry_run: bool = False
    create_lookups: bool = True


@dataclass(frozen=True)
class DbCapabilities:
    has_resource_costs_daily: bool
    resource_costs_daily_columns: Tuple[str, ...]


# PUBLIC_INTERFACE
def discover_input_files(input_dir: Path) -> Tuple[List[Path], List[Path]]:
    """Discover Excel files under input_dir and split into resource files vs recommendation files."""
    xlsx_files = sorted([p for p in input_dir.glob("*.xlsx") if p.is_file()])
    resource_files: List[Path] = []
    recommendation_files: List[Path] = []
    for p in xlsx_files:
        if "recommendation" in p.name.lower():
            recommendation_files.append(p)
        else:
            resource_files.append(p)
    return resource_files, recommendation_files


def _read_excel_rows_openpyxl(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [(_snake(str(c)) if c is not None else "") for c in (rows[0] or [])]
        while headers and headers[-1] == "":
            headers.pop()

        out: List[Dict[str, Any]] = []
        for r in rows[1:]:
            if r is None:
                continue
            values = list(r[: len(headers)])
            row = {headers[i]: values[i] for i in range(len(headers))}
            # skip fully empty rows
            if all(_is_missing(v) for v in row.values()):
                continue
            out.append(row)
        return out
    finally:
        wb.close()


def _read_excel_rows_pandas(path: Path) -> List[Dict[str, Any]]:
    if pd is None:  # pragma: no cover
        return _read_excel_rows_openpyxl(path)

    df = pd.read_excel(path)  # uses openpyxl engine under the hood when installed
    df.columns = [_snake(str(c)) for c in df.columns]
    records: List[Dict[str, Any]] = df.to_dict(orient="records")
    # Drop fully-empty rows
    out: List[Dict[str, Any]] = []
    for r in records:
        if all(_is_missing(v) for v in r.values()):
            continue
        out.append(r)
    return out


def _read_excel_rows(path: Path) -> List[Dict[str, Any]]:
    # Prefer pandas when available; otherwise use openpyxl.
    try:
        if pd is not None:
            return _read_excel_rows_pandas(path)
    except Exception:
        # fallback to openpyxl
        pass
    return _read_excel_rows_openpyxl(path)


# ---------- Upsert SQL (PostgreSQL) ----------

SQL_UPSERT_ORGANIZATION = """
INSERT INTO organizations (id, name, slug)
VALUES (%s, %s, %s)
ON CONFLICT (id) DO UPDATE SET
    name = EXCLUDED.name,
    slug = EXCLUDED.slug,
    updated_at = NOW()
RETURNING (xmax = 0) AS inserted;
"""

SQL_SELECT_ORG_BY_SLUG = """
SELECT id FROM organizations WHERE slug = %s LIMIT 1;
"""

SQL_SELECT_ORG_EXISTS = """
SELECT 1 FROM organizations WHERE id = %s LIMIT 1;
"""

SQL_UPSERT_CLOUD_ACCOUNT_BY_ID = """
INSERT INTO cloud_accounts (id, organization_id, provider, account_id, account_name, connection_config, is_active)
VALUES (%s, %s, %s, %s, %s, %s, TRUE)
ON CONFLICT (id) DO UPDATE SET
    organization_id = EXCLUDED.organization_id,
    provider = EXCLUDED.provider,
    account_id = EXCLUDED.account_id,
    account_name = EXCLUDED.account_name,
    connection_config = EXCLUDED.connection_config,
    is_active = TRUE,
    updated_at = NOW()
RETURNING (xmax = 0) AS inserted;
"""

SQL_SELECT_CLOUD_ACCOUNT_BY_ORG_PROVIDER = """
SELECT id FROM cloud_accounts
WHERE organization_id = %s AND provider = %s AND is_active = TRUE
ORDER BY created_at ASC;
"""

SQL_SELECT_CLOUD_ACCOUNT_BY_ORG_PROVIDER_ACCOUNT_ID = """
SELECT id FROM cloud_accounts
WHERE organization_id = %s AND provider = %s AND account_id = %s
LIMIT 1;
"""

SQL_UPSERT_RESOURCE = """
INSERT INTO resources (
    id, organization_id, cloud_account_id, provider, resource_id,
    resource_type, region, state, tags, cost_daily, cost_monthly, created_at
) VALUES (
    %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s
)
ON CONFLICT (organization_id, provider, resource_id) DO UPDATE SET
    cloud_account_id = EXCLUDED.cloud_account_id,
    resource_type = EXCLUDED.resource_type,
    region = EXCLUDED.region,
    state = EXCLUDED.state,
    tags = EXCLUDED.tags,
    cost_daily = EXCLUDED.cost_daily,
    cost_monthly = EXCLUDED.cost_monthly,
    updated_at = NOW()
RETURNING id, (xmax = 0) AS inserted;
"""

SQL_FIND_RESOURCE_INTERNAL_ID = """
SELECT id FROM resources
WHERE organization_id = %s AND provider = %s AND resource_id = %s
LIMIT 1;
"""

SQL_UPSERT_RECOMMENDATION = """
INSERT INTO recommendations (
    id, organization_id, resource_id, recommendation_type, priority,
    potential_savings_monthly, description, action_items, created_at
) VALUES (
    %s, %s, %s, %s, %s,
    %s, %s, %s, %s
)
ON CONFLICT (id) DO UPDATE SET
    organization_id = EXCLUDED.organization_id,
    resource_id = EXCLUDED.resource_id,
    recommendation_type = EXCLUDED.recommendation_type,
    priority = EXCLUDED.priority,
    potential_savings_monthly = EXCLUDED.potential_savings_monthly,
    description = EXCLUDED.description,
    action_items = EXCLUDED.action_items,
    updated_at = NOW()
RETURNING (xmax = 0) AS inserted;
"""


def _ensure_organization(
    conn: psycopg.Connection,
    org_id: Optional[str],
    org_name: Optional[str],
    org_slug: Optional[str],
    create_lookups: bool,
) -> Tuple[str, bool]:
    """
    Ensure an organization exists. Returns (org_id, created?).
    """
    name = (org_name or "Default Organization").strip()
    slug = _slugify(org_slug or name)

    with conn.cursor(row_factory=dict_row) as cur:
        if org_id:
            cur.execute(SQL_SELECT_ORG_EXISTS, (org_id,))
            exists = cur.fetchone() is not None
            if exists:
                return org_id, False
            if not create_lookups:
                raise SystemExit(f"ERROR: organization {org_id} does not exist; pass --create-lookups to auto-create.")

            # If slug exists already, prefer existing org instead of creating a conflicting duplicate
            cur.execute(SQL_SELECT_ORG_BY_SLUG, (slug,))
            by_slug = cur.fetchone()
            if by_slug:
                return by_slug["id"], False

            cur.execute(SQL_UPSERT_ORGANIZATION, (org_id, name, slug))
            inserted_row = cur.fetchone()
            conn.commit()
            return org_id, bool(inserted_row and inserted_row["inserted"])

        # No org_id provided: try auto detect
        cur.execute("SELECT id FROM organizations ORDER BY created_at ASC")
        rows = cur.fetchall()
        if len(rows) == 1:
            detected = rows[0]["id"]
            logger.info("Auto-detected single organization: %s", detected)
            return detected, False
        if len(rows) > 1:
            raise SystemExit("ERROR: --org/--organization-id is required (multiple organizations exist).")

        # None exist: create if allowed
        if not create_lookups:
            raise SystemExit("ERROR: --org/--organization-id is required (no organizations exist).")

        new_id = str(uuid.uuid5(NAMESPACE_ORG, slug))
        cur.execute(SQL_UPSERT_ORGANIZATION, (new_id, name, slug))
        inserted_row = cur.fetchone()
        conn.commit()
        logger.info("Created organization %s (slug=%s).", new_id, slug)
        return new_id, bool(inserted_row and inserted_row["inserted"])


def _ensure_cloud_account_for_provider(
    conn: psycopg.Connection,
    organization_id: str,
    provider: str,
    desired_cloud_account_id: Optional[str],
    account_native_id: Optional[str],
    account_name: Optional[str],
    create_lookups: bool,
) -> Tuple[str, bool]:
    """
    Ensure a cloud account exists for (organization_id, provider). Returns (cloud_account_id, created?).
    """
    if provider not in PROVIDER_SLUGS:
        raise SystemExit(f"ERROR: invalid provider '{provider}' while ensuring cloud account")

    native_id = None if _is_missing(account_native_id) else str(account_native_id).strip()
    # To keep creation idempotent and avoid NULL in unique keys, provide a stable native_id if none was supplied.
    if native_id is None:
        native_id = f"mock-{provider}"

    name = (account_name or f"{provider.upper()} Account").strip()

    with conn.cursor(row_factory=dict_row) as cur:
        if desired_cloud_account_id:
            # Ensure account row exists by PK
            cur.execute("SELECT 1 FROM cloud_accounts WHERE id=%s LIMIT 1", (desired_cloud_account_id,))
            exists = cur.fetchone() is not None
            if exists:
                return desired_cloud_account_id, False
            if not create_lookups:
                raise SystemExit(
                    f"ERROR: cloud_account_id {desired_cloud_account_id} not found; pass --create-lookups to auto-create."
                )
            cur.execute(
                SQL_UPSERT_CLOUD_ACCOUNT_BY_ID,
                (desired_cloud_account_id, organization_id, provider, native_id, name, Json({})),
            )
            inserted_row = cur.fetchone()
            conn.commit()
            logger.info("Created cloud account %s for provider=%s.", desired_cloud_account_id, provider)
            return desired_cloud_account_id, bool(inserted_row and inserted_row["inserted"])

        # Prefer existing exact match by (org, provider, account_id) when possible
        cur.execute(SQL_SELECT_CLOUD_ACCOUNT_BY_ORG_PROVIDER_ACCOUNT_ID, (organization_id, provider, native_id))
        exact = cur.fetchone()
        if exact:
            return exact["id"], False

        # Otherwise, allow auto-detect if exactly one active exists
        cur.execute(SQL_SELECT_CLOUD_ACCOUNT_BY_ORG_PROVIDER, (organization_id, provider))
        rows = cur.fetchall()
        if len(rows) == 1:
            detected = rows[0]["id"]
            logger.info("Auto-detected cloud account for %s: %s", provider, detected)
            return detected, False
        if len(rows) > 1 and not create_lookups:
            raise SystemExit(
                f"ERROR: multiple active cloud accounts for provider '{provider}'. Pass --account-{provider} or --cloud-account-id."
            )

        if len(rows) == 0 and not create_lookups:
            raise SystemExit(
                f"ERROR: Cloud account for provider '{provider}' required. Pass --account-{provider} or --cloud-account-id."
            )

        # Create stable account ID so reruns are idempotent
        stable_id = str(uuid.uuid5(NAMESPACE_CLOUD_ACCOUNT, f"{organization_id}:{provider}:{native_id}"))
        cur.execute(SQL_UPSERT_CLOUD_ACCOUNT_BY_ID, (stable_id, organization_id, provider, native_id, name, Json({})))
        inserted_row = cur.fetchone()
        conn.commit()
        logger.info("Created cloud account %s for provider=%s (account_id=%s).", stable_id, provider, native_id)
        return stable_id, bool(inserted_row and inserted_row["inserted"])


def _resolve_org_and_accounts(
    conn: psycopg.Connection, args: argparse.Namespace, providers_needed: Iterable[str]
) -> Tuple[str, Dict[str, str], int, int]:
    """
    Resolve (and optionally create) organization and cloud accounts.

    Returns:
        (org_id, account_by_provider, organizations_created_count, cloud_accounts_created_count)
    """
    org_id, org_created = _ensure_organization(
        conn=conn,
        org_id=args.organization_id,
        org_name=args.organization_name,
        org_slug=args.organization_slug,
        create_lookups=bool(args.create_lookups),
    )

    accounts: Dict[str, str] = {}
    cloud_accounts_created = 0

    # Provider-specific desired IDs: --account-aws/--account-azure/--account-gcp
    explicit_id_map = {
        "aws": args.account_aws or args.cloud_account_id,
        "azure": args.account_azure or args.cloud_account_id,
        "gcp": args.account_gcp or args.cloud_account_id,
    }

    # Provider-specific native IDs: --native-account-id-aws/...
    native_id_map = {
        "aws": args.native_account_id_aws,
        "azure": args.native_account_id_azure,
        "gcp": args.native_account_id_gcp,
    }

    # Provider-specific names: --account-name-aws/...
    name_map = {
        "aws": args.account_name_aws,
        "azure": args.account_name_azure,
        "gcp": args.account_name_gcp,
    }

    for prov in providers_needed:
        acc_id, created = _ensure_cloud_account_for_provider(
            conn=conn,
            organization_id=org_id,
            provider=prov,
            desired_cloud_account_id=explicit_id_map.get(prov),
            account_native_id=native_id_map.get(prov),
            account_name=name_map.get(prov),
            create_lookups=bool(args.create_lookups),
        )
        accounts[prov] = acc_id
        if created:
            cloud_accounts_created += 1

    return org_id, accounts, (1 if org_created else 0), cloud_accounts_created


def _resource_cost_daily_row_payload(
    cfg: IngestConfig,
    internal_resource_id: str,
    cloud_account_id: str,
    provider: str,
    region: str,
    cost_daily: Decimal,
    cost_dt: datetime,
    columns: Sequence[str],
) -> Dict[str, Any]:
    """
    Build a flexible payload for resource_costs_daily using only columns that exist in the target DB table.
    """
    cost_day: date = cost_dt.date()
    stable_id = str(uuid.uuid5(NAMESPACE_RESOURCE_COST_DAILY, f"{cfg.organization_id}:{internal_resource_id}:{cost_day.isoformat()}"))

    payload: Dict[str, Any] = {}

    cols = set(columns)

    if "id" in cols:
        payload["id"] = stable_id
    if "organization_id" in cols:
        payload["organization_id"] = cfg.organization_id
    if "cloud_account_id" in cols:
        payload["cloud_account_id"] = cloud_account_id
    if "provider" in cols:
        payload["provider"] = provider
    if "region" in cols:
        payload["region"] = region
    if "resource_id" in cols:
        # IMPORTANT: in our mapping, this is the internal resources.id
        payload["resource_id"] = internal_resource_id
    if "cost_date" in cols:
        payload["cost_date"] = cost_day

    # Common cost column names
    if "cost_amount" in cols:
        payload["cost_amount"] = cost_daily
    if "cost_daily" in cols:
        payload["cost_daily"] = cost_daily
    if "amount" in cols and "cost_amount" not in payload:
        payload["amount"] = cost_daily

    if "currency" in cols:
        payload["currency"] = "USD"

    # Optional timestamps
    if "created_at" in cols:
        payload["created_at"] = cost_dt
    if "updated_at" in cols:
        # let DB default/onupdate handle when possible; on explicit update we may set it
        payload["updated_at"] = cost_dt

    return payload


def _upsert_resource_costs_daily(
    cur: psycopg.Cursor,
    table_name: str,
    columns: Sequence[str],
    payload: Dict[str, Any],
) -> Optional[bool]:
    """
    Upsert a resource_costs_daily row.
    Returns:
      True  -> inserted
      False -> updated
      None  -> skipped (insufficient columns)
    """
    cols = list(payload.keys())
    if not cols:
        return None

    # Prefer ON CONFLICT (id) where id exists.
    if "id" in payload:
        insert_cols = cols
        placeholders = ", ".join(["%s"] * len(insert_cols))
        col_list = ", ".join(insert_cols)

        # update everything except id and created_at
        set_cols = [c for c in insert_cols if c not in {"id", "created_at"}]
        if not set_cols:
            return None

        set_expr = ", ".join([f"{c} = EXCLUDED.{c}" for c in set_cols if c != "updated_at"])
        if "updated_at" in insert_cols:
            # ensure updated_at bumps on update
            if set_expr:
                set_expr = f"{set_expr}, updated_at = NOW()"
            else:
                set_expr = "updated_at = NOW()"

        sql = f"""
        INSERT INTO {table_name} ({col_list})
        VALUES ({placeholders})
        ON CONFLICT (id) DO UPDATE SET
            {set_expr}
        RETURNING (xmax = 0) AS inserted;
        """
        cur.execute(sql, tuple(payload[c] for c in insert_cols))
        row = cur.fetchone()
        return bool(row and row[0])

    # Fallback: if no id column, attempt a "natural-key" update+insert using (organization_id, resource_id, cost_date)
    required = {"organization_id", "resource_id", "cost_date"}
    if not required.issubset(set(payload.keys())):
        return None

    # Try UPDATE first
    set_cols = [c for c in payload.keys() if c not in required and c != "created_at"]
    if not set_cols:
        return None
    set_expr = ", ".join([f"{c} = %s" for c in set_cols])
    upd_sql = f"""
    UPDATE {table_name}
    SET {set_expr}
    WHERE organization_id = %s AND resource_id = %s AND cost_date = %s
    """
    params = [payload[c] for c in set_cols] + [payload["organization_id"], payload["resource_id"], payload["cost_date"]]
    cur.execute(upd_sql, tuple(params))
    if cur.rowcount and cur.rowcount > 0:
        return False

    # INSERT if UPDATE affected none
    insert_cols = list(payload.keys())
    placeholders = ", ".join(["%s"] * len(insert_cols))
    ins_sql = f"INSERT INTO {table_name} ({', '.join(insert_cols)}) VALUES ({placeholders})"
    cur.execute(ins_sql, tuple(payload[c] for c in insert_cols))
    return True


def _ingest_resources_file(
    conn: psycopg.Connection,
    cfg: IngestConfig,
    dbcaps: DbCapabilities,
    path: Path,
) -> Dict[str, TableReport]:
    rows = _read_excel_rows(path)
    reports: Dict[str, TableReport] = {
        "resources": TableReport(),
        "resource_costs_daily": TableReport(),
    }

    if not rows:
        logger.info("No rows found in %s", path.name)
        return reports

    provider = _provider_from_filename(path.name)
    if not provider:
        # Try derive from column if present
        for r in rows:
            provider = _normalize_provider(r.get("provider"))
            if provider:
                break
    if not provider:
        logger.warning("Provider not inferred from filename or data for %s; skipping file.", path.name)
        reports["resources"].skipped += len(rows)
        return reports

    cloud_account_id = cfg.account_by_provider.get(provider)
    if not cloud_account_id:
        logger.warning("No cloud account configured for provider=%s; skipping file %s", provider, path.name)
        reports["resources"].skipped += len(rows)
        return reports

    with conn.cursor() as cur:
        for idx, r in enumerate(rows):
            resource_id = r.get("resource_id") or r.get("id") or r.get("name")
            resource_type = r.get("resource_type")
            region = r.get("region")
            state = r.get("state")
            if _is_missing(resource_id) or _is_missing(resource_type) or _is_missing(region) or _is_missing(state):
                reports["resources"].skipped += 1
                logger.debug(
                    "Skipping row %d in %s: missing required fields (resource_id/resource_type/region/state). Row=%s",
                    idx,
                    path.name,
                    r,
                )
                continue

            tags = _parse_tags(r.get("tags"))
            cost_daily = _to_decimal_or_none(r.get("cost_daily"))
            cost_monthly = _to_decimal_or_none(r.get("cost_monthly"))
            created_at = (
                _parse_datetime_or_none(r.get("launch_time"))
                or _parse_datetime_or_none(r.get("created_at"))
                or _now_utc()
            )

            # Generate a new UUID for first insert; on conflict, this ID is ignored
            res_id = str(uuid.uuid4())

            params = (
                res_id,
                cfg.organization_id,
                cloud_account_id,
                provider,
                str(resource_id),
                str(resource_type),
                str(region),
                _normalize_state(state),
                Json(tags),
                cost_daily,
                cost_monthly,
                created_at,
            )

            if cfg.dry_run:
                existing = _fetch_one_scalar(
                    conn,
                    "SELECT 1 FROM resources WHERE organization_id=%s AND provider=%s AND resource_id=%s",
                    (cfg.organization_id, provider, str(resource_id)),
                )
                if existing:
                    reports["resources"].updated += 1
                else:
                    reports["resources"].inserted += 1

                # Simulate costs table behavior if available and cost_daily present
                if dbcaps.has_resource_costs_daily and cost_daily is not None:
                    # We count as inserted (deterministic id) for a dry-run estimate
                    reports["resource_costs_daily"].inserted += 1
                elif cost_daily is not None and not dbcaps.has_resource_costs_daily:
                    reports["resource_costs_daily"].skipped += 1
                continue

            # Upsert resource and capture internal id for FK usage
            cur.execute(SQL_UPSERT_RESOURCE, params)
            ret = cur.fetchone()
            if not ret:
                reports["resources"].skipped += 1
                continue
            internal_resource_id = str(ret[0])
            was_inserted = bool(ret[1])
            if was_inserted:
                reports["resources"].inserted += 1
            else:
                reports["resources"].updated += 1

            # Optional: Upsert into resource_costs_daily if that table exists in the target DB
            if dbcaps.has_resource_costs_daily and cost_daily is not None:
                payload = _resource_cost_daily_row_payload(
                    cfg=cfg,
                    internal_resource_id=internal_resource_id,
                    cloud_account_id=cloud_account_id,
                    provider=provider,
                    region=str(region),
                    cost_daily=cost_daily,
                    cost_dt=created_at,
                    columns=list(dbcaps.resource_costs_daily_columns),
                )
                try:
                    upserted = _upsert_resource_costs_daily(
                        cur=cur,
                        table_name="resource_costs_daily",
                        columns=list(dbcaps.resource_costs_daily_columns),
                        payload=payload,
                    )
                    if upserted is True:
                        reports["resource_costs_daily"].inserted += 1
                    elif upserted is False:
                        reports["resource_costs_daily"].updated += 1
                    else:
                        reports["resource_costs_daily"].skipped += 1
                except Exception as exc:
                    # Do not fail whole ingestion if optional table differs; log and continue
                    reports["resource_costs_daily"].skipped += 1
                    logger.warning(
                        "Failed to upsert resource_costs_daily for %s row=%d (will skip): %s",
                        path.name,
                        idx,
                        exc,
                    )

    if not cfg.dry_run:
        conn.commit()

    logger.info(
        "Resources file %s: resources(inserted=%s updated=%s skipped=%s) resource_costs_daily(inserted=%s updated=%s skipped=%s)",
        path.name,
        reports["resources"].inserted,
        reports["resources"].updated,
        reports["resources"].skipped,
        reports["resource_costs_daily"].inserted,
        reports["resource_costs_daily"].updated,
        reports["resource_costs_daily"].skipped,
    )
    return reports


def _stable_recommendation_id(
    organization_id: str,
    provider: str,
    external_resource_id: Optional[str],
    recommendation_type: str,
    description: Optional[str],
) -> str:
    base = f"{organization_id}:{provider}:{external_resource_id or ''}:{recommendation_type}:{description or ''}"
    return str(uuid.uuid5(NAMESPACE_RECOMMENDATION, base))


def _ingest_recommendations_file(conn: psycopg.Connection, cfg: IngestConfig, path: Path) -> TableReport:
    rows = _read_excel_rows(path)
    report = TableReport()
    if not rows:
        logger.info("No rows found in %s", path.name)
        return report

    with conn.cursor() as cur:
        for idx, r in enumerate(rows):
            provider = _normalize_provider(r.get("cloud_provider") or r.get("provider")) or _provider_from_filename(path.name) or "aws"

            # External recommendation id (preferred). If missing, generate a stable deterministic ID to preserve idempotency.
            rid = r.get("recommendation_id") or r.get("id")
            external_res_id = None if _is_missing(r.get("resource_id")) else str(r.get("resource_id"))
            rec_type = str(r.get("recommendation_type") or r.get("type") or "General")
            desc = None if _is_missing(r.get("description")) else str(r.get("description"))
            if _is_missing(rid):
                rid = _stable_recommendation_id(
                    organization_id=cfg.organization_id,
                    provider=provider,
                    external_resource_id=external_res_id,
                    recommendation_type=rec_type,
                    description=desc,
                )

            # map resource foreign key if we can (internal resources.id)
            internal_resource_id: Optional[str] = None
            if external_res_id:
                cur.execute(SQL_FIND_RESOURCE_INTERNAL_ID, (cfg.organization_id, provider, external_res_id))
                ir = cur.fetchone()
                if ir:
                    internal_resource_id = ir[0]

            priority = _normalize_priority(r.get("priority"))
            potential_savings = _to_decimal_or_none(r.get("potential_savings_monthly"))

            # action_items: accept a column 'action_items' as CSV or fall back to 'impact' as a single action item
            action_items_raw = r.get("action_items")
            if _is_missing(action_items_raw) and not _is_missing(r.get("impact")):
                action_items: Any = [str(r.get("impact"))]
            elif isinstance(action_items_raw, str):
                parts = [p.strip() for p in action_items_raw.replace(";", ",").split(",") if p.strip()]
                action_items = parts
            elif _is_missing(action_items_raw):
                action_items = None
            else:
                action_items = action_items_raw  # could already be a list-like

            params = (
                str(rid),
                cfg.organization_id,
                internal_resource_id,
                rec_type,
                priority,
                potential_savings,
                desc,
                Json(action_items) if action_items is not None else None,
                _now_utc(),
            )

            if cfg.dry_run:
                existing = _fetch_one_scalar(conn, "SELECT 1 FROM recommendations WHERE id=%s", (str(rid),))
                if existing:
                    report.updated += 1
                else:
                    report.inserted += 1
                continue

            cur.execute(SQL_UPSERT_RECOMMENDATION, params)
            ret = cur.fetchone()
            was_inserted = bool(ret and ret[0])
            if was_inserted:
                report.inserted += 1
            else:
                report.updated += 1

    if not cfg.dry_run:
        conn.commit()

    logger.info(
        "Recommendations file %s: inserted=%s updated=%s skipped=%s",
        path.name,
        report.inserted,
        report.updated,
        report.skipped,
    )
    return report


# PUBLIC_INTERFACE
def run() -> int:
    """Entry point for CLI execution."""
    load_dotenv()

    parser = argparse.ArgumentParser(description="CloudUnify Pro dataset ingestion tool")

    parser.add_argument("--database-url", dest="database_url", help="Database URL (postgresql://...).")
    parser.add_argument("--org", "--organization-id", dest="organization_id", help="Organization UUID to assign data to.")
    parser.add_argument("--org-name", "--organization-name", dest="organization_name", help="Organization name (used when creating org).")
    parser.add_argument("--org-slug", "--organization-slug", dest="organization_slug", help="Organization slug (used when creating org).")

    parser.add_argument("--create-lookups", dest="create_lookups", action="store_true", default=True, help="Create organizations/cloud_accounts as needed (default: enabled).")
    parser.add_argument("--no-create-lookups", dest="create_lookups", action="store_false", help="Disable auto-creation of lookups.")

    parser.add_argument("--cloud-account-id", dest="cloud_account_id", help="Cloud account UUID to use for all providers (if provider-specific not provided).")
    parser.add_argument("--account-aws", dest="account_aws", help="Cloud account UUID for AWS provider.")
    parser.add_argument("--account-azure", dest="account_azure", help="Cloud account UUID for Azure provider.")
    parser.add_argument("--account-gcp", dest="account_gcp", help="Cloud account UUID for GCP provider.")

    parser.add_argument("--native-account-id-aws", dest="native_account_id_aws", help="Provider-native account id (AWS account number) when auto-creating AWS cloud account.")
    parser.add_argument("--native-account-id-azure", dest="native_account_id_azure", help="Provider-native account id when auto-creating Azure cloud account.")
    parser.add_argument("--native-account-id-gcp", dest="native_account_id_gcp", help="Provider-native account id when auto-creating GCP cloud account.")

    parser.add_argument("--account-name-aws", dest="account_name_aws", help="Account display name for AWS cloud account (when creating).")
    parser.add_argument("--account-name-azure", dest="account_name_azure", help="Account display name for Azure cloud account (when creating).")
    parser.add_argument("--account-name-gcp", dest="account_name_gcp", help="Account display name for GCP cloud account (when creating).")

    parser.add_argument("--input-dir", dest="input_dir", default="", help="Directory containing .xlsx files. Defaults to repo_root/attachments.")
    parser.add_argument("--dry-run", dest="dry_run", action="store_true", help="Validate and report without writing to DB.")
    args = parser.parse_args()

    # Resolve input directory
    input_dir = Path(args.input_dir) if args.input_dir else None
    if input_dir is None or not input_dir.exists():
        script_dir = Path(__file__).resolve().parent
        repo_root = _resolve_repo_root(script_dir.parent) or script_dir.parent.parent.parent
        input_dir = repo_root / "attachments"
    if not input_dir.exists():
        logger.error("Input directory does not exist: %s", input_dir)
        return 2

    # Resolve DSN with required precedence:
    #   CLI > db_connection.txt > env DATABASE_URL > fallback
    raw_dsn: Optional[str] = args.database_url
    if not raw_dsn:
        dbc_path = _find_db_connection_txt(Path(__file__).resolve().parent)
        if dbc_path:
            try:
                txt = dbc_path.read_text(encoding="utf-8", errors="ignore")
                parsed = _parse_db_connection_txt(txt)
                if parsed:
                    raw_dsn = parsed
                    logger.info("Using DB connection from %s", dbc_path)
            except Exception as exc:
                logger.warning("Failed reading %s (%s); falling back to env DATABASE_URL", dbc_path, exc)

    if not raw_dsn:
        raw_dsn = os.getenv("DATABASE_URL")

    if not raw_dsn:
        raw_dsn = FALLBACK_NEON_DSN

    dsn = _ensure_ssl_and_channel_binding(raw_dsn)

    # Discover files
    resource_files, recommendation_files = discover_input_files(input_dir)
    if not resource_files and not recommendation_files:
        logger.warning("No Excel files found in %s", input_dir)
        return 0

    # Determine providers present in resource files
    providers_in_files: set[str] = set()
    for p in resource_files:
        prov = _provider_from_filename(p.name)
        if prov:
            providers_in_files.add(prov)
    providers_needed = sorted(providers_in_files or list(PROVIDER_SLUGS))

    logger.info("Connecting to database...")
    with psycopg.connect(dsn, autocommit=False) as conn:
        # Detect optional tables
        has_rcd = _table_exists(conn, "resource_costs_daily")
        rcd_cols = tuple(_get_table_columns(conn, "resource_costs_daily")) if has_rcd else tuple()
        dbcaps = DbCapabilities(has_resource_costs_daily=has_rcd, resource_costs_daily_columns=rcd_cols)
        if has_rcd:
            logger.info("Detected table resource_costs_daily (columns=%s)", ", ".join(rcd_cols))
        else:
            logger.info("Table resource_costs_daily not found in target DB; will skip that portion of import.")

        # Resolve/create organization and cloud accounts
        org_id, accounts, orgs_created, accounts_created = _resolve_org_and_accounts(conn, args, providers_needed)
        cfg = IngestConfig(
            organization_id=org_id,
            account_by_provider=accounts,
            dsn=dsn,
            input_dir=input_dir,
            dry_run=bool(args.dry_run),
            create_lookups=bool(args.create_lookups),
        )

        totals: Dict[str, TableReport] = {
            "resources": TableReport(),
            "resource_costs_daily": TableReport(),
            "recommendations": TableReport(),
        }

        # Process resource files first (FKs for recommendations)
        for rf in resource_files:
            try:
                rep = _ingest_resources_file(conn, cfg, dbcaps, rf)
                totals["resources"].inserted += rep["resources"].inserted
                totals["resources"].updated += rep["resources"].updated
                totals["resources"].skipped += rep["resources"].skipped

                totals["resource_costs_daily"].inserted += rep["resource_costs_daily"].inserted
                totals["resource_costs_daily"].updated += rep["resource_costs_daily"].updated
                totals["resource_costs_daily"].skipped += rep["resource_costs_daily"].skipped
            except Exception as exc:
                conn.rollback()
                logger.exception("Failed ingesting resources file %s (rolled back that file): %s", rf.name, exc)

        for recf in recommendation_files:
            try:
                rep = _ingest_recommendations_file(conn, cfg, recf)
                totals["recommendations"].inserted += rep.inserted
                totals["recommendations"].updated += rep.updated
                totals["recommendations"].skipped += rep.skipped
            except Exception as exc:
                conn.rollback()
                logger.exception("Failed ingesting recommendations file %s (rolled back that file): %s", recf.name, exc)

        # Concise import report
        logger.info("==== Import Report ====")
        logger.info("Lookups: organizations_created=%s cloud_accounts_created=%s", orgs_created, accounts_created)
        logger.info(
            "resources: inserted=%s updated=%s skipped=%s",
            totals["resources"].inserted,
            totals["resources"].updated,
            totals["resources"].skipped,
        )
        if dbcaps.has_resource_costs_daily:
            logger.info(
                "resource_costs_daily: inserted=%s updated=%s skipped=%s",
                totals["resource_costs_daily"].inserted,
                totals["resource_costs_daily"].updated,
                totals["resource_costs_daily"].skipped,
            )
        else:
            logger.info("resource_costs_daily: skipped (table not present)")
        logger.info(
            "recommendations: inserted=%s updated=%s skipped=%s",
            totals["recommendations"].inserted,
            totals["recommendations"].updated,
            totals["recommendations"].skipped,
        )

    return 0


if __name__ == "__main__":
    sys.exit(run())
