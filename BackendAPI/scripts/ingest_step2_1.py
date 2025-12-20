#!/usr/bin/env python3
"""
Step 2.1 ingestion runner (Neon/PostgreSQL)

This script executes the step 2.1 ingestion workflow against the *actual* Neon schema
present in this environment:

- organizations(id uuid default gen_random_uuid(), name, slug, created_at)
- cloud_accounts(id uuid default gen_random_uuid(), organization_id, provider, account_external_id, account_name, is_active, created_at)
- resources(id uuid default gen_random_uuid(), organization_id, cloud_account_id, resource_id, resource_type, provider, region, state, tags jsonb, created_at)
  - Unique key: (cloud_account_id, resource_id)
- resource_costs_daily(id bigint, organization_id, resource_id, usage_date, cost_daily, currency)
  - Unique key: (resource_id, usage_date)

It:
- Ensures default organization "CloudUnify Demo" (slug "cloudunify-demo") exists.
- Ensures default cloud accounts exist:
  - AWS Main (aws), Azure Main (azure), GCP Main (gcp)
- Ingests the provided AWS resources XLSX into:
  - resources
  - resource_costs_daily (if cost_daily present; includes 0)

DB connection precedence:
1) --database-url
2) DATABASE_URL env var

Neon requirements:
- Enforces sslmode=require and channel_binding=require.

Run example:
  python scripts/ingest_step2_1.py \
    --database-url "postgresql://.../neondb?sslmode=require&channel_binding=require" \
    --xlsx /path/to/20251220_123808_mock_aws_resources_1.xlsx
"""

from __future__ import annotations

import argparse
import ast
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import psycopg
from openpyxl import load_workbook
from psycopg.types.json import Jsonb


@dataclass
class UpsertCounts:
    inserted: int = 0
    updated: int = 0
    skipped: int = 0


def _ensure_ssl_and_channel_binding(dsn: str) -> str:
    """Ensure DSN includes Neon-required sslmode and channel_binding flags."""
    parts = urlparse(dsn)
    q = dict(parse_qsl(parts.query, keep_blank_values=True))
    q.setdefault("sslmode", "require")
    q.setdefault("channel_binding", "require")
    return urlunparse(parts._replace(query=urlencode(q)))


def _parse_iso_datetime(v: Any) -> Optional[datetime]:
    """Parse ISO-8601 timestamps, including trailing 'Z'."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)

    s = str(v).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _parse_tags_cell(v: Any) -> Dict[str, str]:
    """Parse tags from XLSX. Supports python-dict strings like \"{'k':'v'}\"."""
    if v is None:
        return {}
    if isinstance(v, dict):
        return {str(k): str(val) for k, val in v.items()}

    s = str(v).strip()
    if not s:
        return {}

    if s.startswith("{") and s.endswith("}"):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, dict):
                return {str(k): str(val) for k, val in parsed.items()}
        except Exception:
            pass

    # Fallback "k:v,k2:v2"
    out: Dict[str, str] = {}
    for pair in s.split(","):
        pair = pair.strip()
        if not pair:
            continue
        if ":" in pair:
            k, val = pair.split(":", 1)
            out[k.strip()] = val.strip()
        else:
            out[pair] = "true"
    return out


def _to_decimal(v: Any) -> Optional[Decimal]:
    """Convert numeric-ish values to Decimal; returns None on invalid/blank."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _upsert_org(conn: psycopg.Connection, *, name: str, slug: str) -> Tuple[str, bool]:
    """Upsert organization by slug. Returns (org_id, inserted?)."""
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO organizations (name, slug)
            VALUES (%s, %s)
            ON CONFLICT (slug) DO UPDATE SET
              name = EXCLUDED.name
            RETURNING id, (xmax = 0) AS inserted
            """,
            (name, slug),
        )
        org_id, inserted = cur.fetchone()
        return str(org_id), bool(inserted)


def _upsert_cloud_account(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    provider: str,
    account_external_id: str,
    account_name: str,
) -> Tuple[str, bool]:
    """Upsert cloud account by (organization_id, provider, account_external_id)."""
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO cloud_accounts (organization_id, provider, account_external_id, account_name, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
            ON CONFLICT (organization_id, provider, account_external_id) DO UPDATE SET
              account_name = EXCLUDED.account_name,
              is_active = TRUE
            RETURNING id, (xmax = 0) AS inserted
            """,
            (organization_id, provider, account_external_id, account_name),
        )
        account_id, inserted = cur.fetchone()
        return str(account_id), bool(inserted)


def _ingest_aws_xlsx(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    cloud_account_id: str,
    provider: str,
    xlsx_path: Path,
    populate_daily_costs: bool,
) -> Tuple[UpsertCounts, UpsertCounts]:
    """Ingest AWS resources XLSX into resources and (optionally) resource_costs_daily."""
    resources_counts = UpsertCounts()
    rcd_counts = UpsertCounts()

    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_row = next(rows_iter)
        except StopIteration:
            return resources_counts, rcd_counts

        headers = [str(h).strip() if h is not None else "" for h in headers_row]
        index = {h: i for i, h in enumerate(headers)}

        required = ["resource_id", "resource_type", "region", "state"]
        for req in required:
            if req not in index:
                raise RuntimeError(f"XLSX missing required column: {req}")

        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue

            resource_id = row[index["resource_id"]]
            resource_type = row[index["resource_type"]]
            region = row[index["region"]]
            state = row[index["state"]]

            if resource_id is None or resource_type is None:
                resources_counts.skipped += 1
                continue

            launch_time = row[index.get("launch_time", -1)] if "launch_time" in index else None
            created_at = _parse_iso_datetime(launch_time) or datetime.now(timezone.utc)

            tags_raw = row[index.get("tags", -1)] if "tags" in index else None
            tags = _parse_tags_cell(tags_raw)

            # Upsert resource (conflict on unique key (cloud_account_id, resource_id))
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO resources (
                      organization_id, cloud_account_id, resource_id, resource_type,
                      provider, region, state, tags, created_at
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (cloud_account_id, resource_id) DO UPDATE SET
                      organization_id = EXCLUDED.organization_id,
                      resource_type = EXCLUDED.resource_type,
                      provider = EXCLUDED.provider,
                      region = EXCLUDED.region,
                      state = EXCLUDED.state,
                      tags = EXCLUDED.tags
                    RETURNING id, (xmax = 0) AS inserted
                    """,
                    (
                        organization_id,
                        cloud_account_id,
                        str(resource_id),
                        str(resource_type),
                        provider,
                        str(region) if region is not None else None,
                        str(state).strip().lower() if state is not None else None,
                        Jsonb(tags),
                        created_at,
                    ),
                )
                internal_resource_id, inserted = cur.fetchone()

            if inserted:
                resources_counts.inserted += 1
            else:
                resources_counts.updated += 1

            # Optional: daily costs (resource_costs_daily)
            if populate_daily_costs and "cost_daily" in index:
                cost_daily_val = _to_decimal(row[index["cost_daily"]])
                if cost_daily_val is not None:
                    usage_date = created_at.date()
                    with conn.cursor() as cur:
                        cur.execute(
                            """
                            INSERT INTO resource_costs_daily (
                              organization_id, resource_id, usage_date, cost_daily, currency
                            )
                            VALUES (%s,%s,%s,%s,%s)
                            ON CONFLICT (resource_id, usage_date) DO UPDATE SET
                              organization_id = EXCLUDED.organization_id,
                              cost_daily = EXCLUDED.cost_daily,
                              currency = EXCLUDED.currency
                            RETURNING id, (xmax = 0) AS inserted
                            """,
                            (
                                organization_id,
                                str(internal_resource_id),
                                usage_date,
                                cost_daily_val,
                                "USD",
                            ),
                        )
                        _rcd_id, rcd_inserted = cur.fetchone()
                    if rcd_inserted:
                        rcd_counts.inserted += 1
                    else:
                        rcd_counts.updated += 1

        return resources_counts, rcd_counts
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="CloudUnify Pro step 2.1 ingestion runner (Neon)")
    parser.add_argument("--database-url", dest="database_url", default="", help="Neon DATABASE_URL (postgresql://...)")
    parser.add_argument(
        "--xlsx",
        dest="xlsx",
        default="/home/kavia/workspace/code-generation/attachments/20251220_123808_mock_aws_resources_1.xlsx",
        help="Path to the AWS resources XLSX for step 2.1.",
    )
    parser.add_argument("--org-name", dest="org_name", default="CloudUnify Demo")
    parser.add_argument("--org-slug", dest="org_slug", default="cloudunify-demo")
    parser.add_argument("--populate-daily-costs", dest="populate_daily_costs", action="store_true", default=True)
    args = parser.parse_args()

    raw_dsn = args.database_url or os.getenv("DATABASE_URL", "")
    if not raw_dsn:
        print("ERROR: provide --database-url or set DATABASE_URL", file=sys.stderr)
        return 2

    dsn = _ensure_ssl_and_channel_binding(raw_dsn)
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: XLSX not found: {xlsx_path}", file=sys.stderr)
        return 2

    with psycopg.connect(dsn, autocommit=False) as conn:
        # 1) Ensure org
        org_id, org_inserted = _upsert_org(conn, name=args.org_name, slug=args.org_slug)

        # 2) Ensure cloud accounts (defaults)
        accounts_inserted = 0
        accounts: Dict[str, str] = {}

        for provider, account_name in (("aws", "AWS Main"), ("azure", "Azure Main"), ("gcp", "GCP Main")):
            account_id, inserted = _upsert_cloud_account(
                conn,
                organization_id=org_id,
                provider=provider,
                account_external_id=f"mock-{provider}",
                account_name=account_name,
            )
            accounts[provider] = account_id
            accounts_inserted += 1 if inserted else 0

        # 3) Ingest AWS XLSX into resources + resource_costs_daily
        res_counts, rcd_counts = _ingest_aws_xlsx(
            conn,
            organization_id=org_id,
            cloud_account_id=accounts["aws"],
            provider="aws",
            xlsx_path=xlsx_path,
            populate_daily_costs=bool(args.populate_daily_costs),
        )

        conn.commit()

    # Concise report
    print("==== Step 2.1 Import Report ====")
    print(f"organization: name={args.org_name!r} slug={args.org_slug!r} id={org_id} created={org_inserted}")
    print(f"cloud_accounts_created={accounts_inserted} cloud_accounts_by_provider={accounts}")
    print(f"file={xlsx_path}")
    print(f"resources: inserted={res_counts.inserted} updated={res_counts.updated} skipped={res_counts.skipped}")
    print(
        "resource_costs_daily: "
        f"inserted={rcd_counts.inserted} updated={rcd_counts.updated} skipped={rcd_counts.skipped}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
