#!/usr/bin/env python3
"""
CloudUnify Pro - Step 2.2 → 2.5 ingestion runner (Neon/PostgreSQL)

This script ingests the following XLSX attachments into the *actual Neon schema*
present in this environment:

- Step 2.2: AWS resources XLSX  -> resources + resource_costs_daily (if present)
- Step 2.3: Azure resources XLSX -> resources + resource_costs_daily (if present)
- Step 2.4: GCP resources XLSX  -> resources + resource_costs_daily (if present)
- Step 2.5: Recommendations XLSX -> recommendations (linked to resources by stable keys)

IMPORTANT: Neon schema differences vs the app's ORM/migrations
- cloud_accounts uses `account_external_id` (NOT `account_id`).
- resources unique constraint is (cloud_account_id, resource_id) and the table
  does NOT include cost_daily/cost_monthly columns (daily cost lives in resource_costs_daily).
- resource_costs_daily unique constraint is (resource_id, usage_date) and `id` is bigint PK.
- recommendations.action_items is `text[]` (ARRAY), not JSON.

Defaults (as requested by the work item)
- Organization: "CloudUnify Demo" slug "cloudunify-demo"
- Cloud accounts (account_external_id = mock-<provider>):
  - AWS:   "AWS Main"
  - Azure: "Azure Main"
  - GCP:   "GCP Main"

DB connection
- Prefer --database-url, else DATABASE_URL environment variable.
- DSN must include sslmode=require and channel_binding=require (Neon). The script enforces these.

Outputs
- Writes a consolidated final import report to:
  - logs/import_report_final_20251220.md
  - logs/import_report_final_20251220.json (optional; enabled by default)

Run (example)
  cd cloudunify-pro-v1-287143-293501/BackendAPI
  python scripts/ingest_step2_2_to_2_5.py \
    --database-url "postgresql://.../neondb?sslmode=require&channel_binding=require" \
    --report-md /home/kavia/workspace/code-generation/logs/import_report_final_20251220.md \
    --report-json /home/kavia/workspace/code-generation/logs/import_report_final_20251220.json
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import uuid
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import psycopg
from openpyxl import load_workbook
from psycopg.types.json import Jsonb


# Stable UUID namespace for deterministic recommendation IDs (keeps reruns idempotent)
NAMESPACE_RECOMMENDATION = uuid.UUID("1d4ab8e8-0d5e-4c1a-9a41-2a30f1a34a2c")

PROVIDERS = ("aws", "azure", "gcp")
PRIORITIES = {"low", "medium", "high", "critical"}


@dataclass
class TableCounts:
    """Counts for a single target table."""

    inserted: int = 0
    updated: int = 0
    skipped: int = 0


@dataclass
class FileReport:
    """Per-file ingestion results."""

    file: str
    provider: Optional[str]
    tables: Dict[str, TableCounts]


def _ensure_ssl_and_channel_binding(dsn: str) -> str:
    """Ensure DSN has sslmode=require and channel_binding=require."""
    parts = urlparse(dsn)
    q = dict(parse_qsl(parts.query, keep_blank_values=True))
    q.setdefault("sslmode", "require")
    q.setdefault("channel_binding", "require")
    return urlunparse(parts._replace(query=urlencode(q)))


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _is_missing(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    if isinstance(v, float) and math.isnan(v):
        return True
    return False


def _snake(s: str) -> str:
    return s.strip().lower().replace(" ", "_")


def _provider_from_filename(name: str) -> Optional[str]:
    lower = name.lower()
    for prov in PROVIDERS:
        if prov in lower:
            return prov
    return None


def _normalize_provider(v: Any) -> Optional[str]:
    """Normalize provider strings like 'AWS'/'Azure' into 'aws'/'azure'."""
    if _is_missing(v):
        return None
    s = str(v).strip().lower()
    if s in {"aws", "amazon", "amazon web services"}:
        return "aws"
    if s in {"azure", "microsoft", "microsoft azure"}:
        return "azure"
    if s in {"gcp", "google", "google cloud"}:
        return "gcp"
    if s in PROVIDERS:
        return s
    return None


def _normalize_priority(v: Any) -> str:
    if _is_missing(v):
        return "medium"
    s = str(v).strip().lower()
    return s if s in PRIORITIES else "medium"


def _parse_datetime(v: Any) -> Optional[datetime]:
    """Parse datetime from openpyxl cell values or ISO-like strings."""
    if _is_missing(v):
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _to_decimal_or_none(v: Any) -> Optional[Decimal]:
    if _is_missing(v):
        return None
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v).strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_tags(v: Any) -> Dict[str, str]:
    """Parse tags from either dict or 'k:v,k2:v2' string."""
    if _is_missing(v):
        return {}
    if isinstance(v, dict):
        return {str(k): str(val) for k, val in v.items() if not _is_missing(k)}
    out: Dict[str, str] = {}
    for pair in str(v).split(","):
        pair = pair.strip()
        if not pair:
            continue
        if ":" in pair:
            k, val = pair.split(":", 1)
            out[k.strip()] = val.strip()
        else:
            out[pair] = "true"
    return out


def _read_xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    """Read XLSX rows as a list of dicts using normalized snake_case headers."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_snake(str(h)) if h is not None else "" for h in rows[0]]
        while headers and headers[-1] == "":
            headers.pop()

        out: List[Dict[str, Any]] = []
        for r in rows[1:]:
            if r is None:
                continue
            values = list(r[: len(headers)])
            row = {headers[i]: values[i] for i in range(len(headers))}
            if all(_is_missing(v) for v in row.values()):
                continue
            out.append(row)
        return out
    finally:
        wb.close()


def _table_exists(conn: psycopg.Connection, table: str) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema='public' AND table_name=%s
            LIMIT 1
            """,
            (table,),
        )
        return cur.fetchone() is not None


def _upsert_org(conn: psycopg.Connection, name: str, slug: str) -> Tuple[str, bool]:
    """Upsert organization by slug. Returns (org_id, inserted?)."""
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO organizations (name, slug)
            VALUES (%s, %s)
            ON CONFLICT (slug) DO UPDATE SET
              name = EXCLUDED.name
            RETURNING id, (xmax = 0) AS inserted
            """,
            (name, slug),
        )
        org_id, inserted = cur.fetchone()
        return str(org_id), bool(inserted)


def _upsert_cloud_account(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    provider: str,
    account_external_id: str,
    account_name: str,
) -> Tuple[str, bool]:
    """Upsert cloud account by (organization_id, provider, account_external_id). Returns (id, inserted?)."""
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO cloud_accounts (organization_id, provider, account_external_id, account_name, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
            ON CONFLICT (organization_id, provider, account_external_id) DO UPDATE SET
              account_name = EXCLUDED.account_name,
              is_active = TRUE
            RETURNING id, (xmax = 0) AS inserted
            """,
            (organization_id, provider, account_external_id, account_name),
        )
        account_id, inserted = cur.fetchone()
        return str(account_id), bool(inserted)


def _stable_recommendation_uuid(
    organization_id: str, provider: str, recommendation_id: Any, row: Dict[str, Any]
) -> str:
    """
    Generate a stable UUID for recommendations.

    If recommendation_id is already a UUID, use it.
    Otherwise derive UUIDv5 from a stable base string, including provider + external recommendation_id.
    """
    if not _is_missing(recommendation_id):
        try:
            return str(uuid.UUID(str(recommendation_id)))
        except Exception:
            pass

    base = (
        f"{organization_id}:{provider}:"
        f"{row.get('recommendation_id') or ''}:"
        f"{row.get('resource_id') or ''}:"
        f"{row.get('recommendation_type') or ''}:"
        f"{row.get('description') or ''}"
    )
    # Avoid very long v5 names; strip repeated whitespace
    base = re.sub(r"\s+", " ", base).strip()
    return str(uuid.uuid5(NAMESPACE_RECOMMENDATION, base))


def _find_resource_internal_id(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    provider: str,
    cloud_account_id: str,
    external_resource_id: str,
) -> Optional[str]:
    """
    Resolve a resource internal UUID by stable key.

    We key by (cloud_account_id, resource_id) which matches Neon uniqueness.
    Also filter by provider + organization_id for safety.
    """
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT id
            FROM resources
            WHERE organization_id=%s
              AND provider=%s
              AND cloud_account_id=%s
              AND resource_id=%s
            ORDER BY created_at ASC
            LIMIT 1
            """,
            (organization_id, provider, cloud_account_id, external_resource_id),
        )
        row = cur.fetchone()
        return str(row[0]) if row else None


def _ingest_resources_xlsx(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    provider: str,
    cloud_account_id: str,
    path: Path,
    has_resource_costs_daily: bool,
) -> FileReport:
    rows = _read_xlsx_rows(path)
    tables = {"resources": TableCounts(), "resource_costs_daily": TableCounts()}

    if not rows:
        return FileReport(file=str(path), provider=provider, tables=tables)

    with conn.cursor() as cur:
        for r in rows:
            external_id = r.get("resource_id")
            resource_type = r.get("resource_type")
            # GCP file has both region + zone; use region if present, else zone.
            region = r.get("region") or r.get("zone")
            state = r.get("state")

            if any(_is_missing(v) for v in (external_id, resource_type, region, state)):
                tables["resources"].skipped += 1
                tables["resource_costs_daily"].skipped += 1
                continue

            created_at = _parse_datetime(r.get("launch_time")) or _now_utc()
            tags = _parse_tags(r.get("tags"))

            cur.execute(
                """
                INSERT INTO resources (
                  organization_id, cloud_account_id, resource_id, resource_type,
                  provider, region, state, tags, created_at
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (cloud_account_id, resource_id) DO UPDATE SET
                  organization_id = EXCLUDED.organization_id,
                  resource_type = EXCLUDED.resource_type,
                  provider = EXCLUDED.provider,
                  region = EXCLUDED.region,
                  state = EXCLUDED.state,
                  tags = EXCLUDED.tags
                RETURNING id, (xmax = 0) AS inserted
                """,
                (
                    organization_id,
                    cloud_account_id,
                    str(external_id),
                    str(resource_type),
                    provider,
                    str(region),
                    str(state).strip().lower(),
                    Jsonb(tags),
                    created_at,
                ),
            )
            internal_id, inserted = cur.fetchone()
            if inserted:
                tables["resources"].inserted += 1
            else:
                tables["resources"].updated += 1

            # Costs: only if the table exists, and the XLSX provides a cost_daily value.
            if not has_resource_costs_daily:
                tables["resource_costs_daily"].skipped += 1
                continue

            cost_daily = _to_decimal_or_none(r.get("cost_daily"))
            if cost_daily is None:
                tables["resource_costs_daily"].skipped += 1
                continue

            usage_date: date = created_at.date()
            cur.execute(
                """
                INSERT INTO resource_costs_daily (organization_id, resource_id, usage_date, cost_daily, currency)
                VALUES (%s,%s,%s,%s,%s)
                ON CONFLICT (resource_id, usage_date) DO UPDATE SET
                  organization_id = EXCLUDED.organization_id,
                  cost_daily = EXCLUDED.cost_daily,
                  currency = EXCLUDED.currency
                RETURNING (xmax = 0) AS inserted
                """,
                (organization_id, str(internal_id), usage_date, cost_daily, "USD"),
            )
            (c_inserted,) = cur.fetchone()
            if c_inserted:
                tables["resource_costs_daily"].inserted += 1
            else:
                tables["resource_costs_daily"].updated += 1

    conn.commit()
    return FileReport(file=str(path), provider=provider, tables=tables)


def _ingest_recommendations_xlsx(
    conn: psycopg.Connection,
    *,
    organization_id: str,
    cloud_account_ids: Dict[str, str],
    path: Path,
) -> FileReport:
    rows = _read_xlsx_rows(path)
    tables = {"recommendations": TableCounts()}

    if not rows:
        return FileReport(file=str(path), provider=None, tables=tables)

    with conn.cursor() as cur:
        for r in rows:
            provider = _normalize_provider(r.get("cloud_provider")) or _provider_from_filename(path.name) or "aws"
            rec_uuid = _stable_recommendation_uuid(organization_id, provider, r.get("recommendation_id"), r)

            # Link to resource internal id using stable keys when possible
            external_res_id = None if _is_missing(r.get("resource_id")) else str(r.get("resource_id"))
            internal_res_id: Optional[str] = None
            if external_res_id and provider in cloud_account_ids:
                internal_res_id = _find_resource_internal_id(
                    conn,
                    organization_id=organization_id,
                    provider=provider,
                    cloud_account_id=cloud_account_ids[provider],
                    external_resource_id=external_res_id,
                )

            rec_type = str(r.get("recommendation_type") or "General")
            priority = _normalize_priority(r.get("priority"))
            savings = _to_decimal_or_none(r.get("potential_savings_monthly"))
            description = None if _is_missing(r.get("description")) else str(r.get("description"))

            # recommendations.action_items is text[] in Neon; store impact as a single action item
            action_items: Optional[List[str]] = None
            if not _is_missing(r.get("impact")):
                action_items = [str(r.get("impact"))]

            cur.execute(
                """
                INSERT INTO recommendations (
                  id, organization_id, resource_id, recommendation_type, priority,
                  potential_savings_monthly, description, action_items, created_at
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id) DO UPDATE SET
                  organization_id = EXCLUDED.organization_id,
                  resource_id = EXCLUDED.resource_id,
                  recommendation_type = EXCLUDED.recommendation_type,
                  priority = EXCLUDED.priority,
                  potential_savings_monthly = EXCLUDED.potential_savings_monthly,
                  description = EXCLUDED.description,
                  action_items = EXCLUDED.action_items
                RETURNING (xmax = 0) AS inserted
                """,
                (
                    rec_uuid,
                    organization_id,
                    internal_res_id,
                    rec_type,
                    priority,
                    savings,
                    description,
                    action_items,
                    _now_utc(),
                ),
            )
            (inserted,) = cur.fetchone()
            if inserted:
                tables["recommendations"].inserted += 1
            else:
                tables["recommendations"].updated += 1

    conn.commit()
    return FileReport(file=str(path), provider=None, tables=tables)


def _sum_table_counts(reports: Iterable[FileReport]) -> Dict[str, TableCounts]:
    totals: Dict[str, TableCounts] = {}
    for fr in reports:
        for table, cnt in fr.tables.items():
            if table not in totals:
                totals[table] = TableCounts()
            totals[table].inserted += cnt.inserted
            totals[table].updated += cnt.updated
            totals[table].skipped += cnt.skipped
    return totals


def _render_markdown_report(
    *,
    started_at: datetime,
    finished_at: datetime,
    database_url: str,
    org: Dict[str, Any],
    lookups: Dict[str, TableCounts],
    file_reports: List[FileReport],
    totals: Dict[str, TableCounts],
) -> str:
    def _row(table: str, c: TableCounts) -> str:
        return f"| {table} | {c.inserted} | {c.updated} | {c.skipped} |"

    # Redact credentials in DSN for markdown (keep host/db).
    safe_dsn = database_url
    try:
        parsed = urlparse(database_url)
        safe_netloc = parsed.hostname or ""
        if parsed.port:
            safe_netloc += f":{parsed.port}"
        safe_dsn = urlunparse(parsed._replace(netloc=safe_netloc))
    except Exception:
        pass

    lines: List[str] = []
    lines.append("# CloudUnify Pro — Final Import Report (2025-12-20)")
    lines.append("")
    lines.append("## Run metadata")
    lines.append("")
    lines.append(f"- Started: `{started_at.isoformat()}`")
    lines.append(f"- Finished: `{finished_at.isoformat()}`")
    lines.append(f"- Database: `{safe_dsn}`")
    lines.append("")
    lines.append("## Defaults used")
    lines.append("")
    lines.append(f"- Organization: **{org.get('name')}** (slug: `{org.get('slug')}`, id: `{org.get('id')}`)")
    lines.append("- Cloud Accounts:")
    for prov in PROVIDERS:
        ca = org.get("cloud_accounts", {}).get(prov, {})
        lines.append(f"  - {prov}: **{ca.get('name')}** (id: `{ca.get('id')}`, account_external_id: `{ca.get('account_external_id')}`)")
    lines.append("")
    lines.append("## Lookup upserts")
    lines.append("")
    lines.append("| table | inserted | updated | skipped |")
    lines.append("|---|---:|---:|---:|")
    for t in ("organizations", "cloud_accounts"):
        c = lookups.get(t, TableCounts())
        lines.append(_row(t, c))
    lines.append("")
    lines.append("## Per-file results")
    lines.append("")
    for fr in file_reports:
        lines.append(f"### {Path(fr.file).name}")
        if fr.provider:
            lines.append(f"- Provider: `{fr.provider}`")
        lines.append("")
        lines.append("| table | inserted | updated | skipped |")
        lines.append("|---|---:|---:|---:|")
        for table_name, cnt in fr.tables.items():
            lines.append(_row(table_name, cnt))
        lines.append("")
    lines.append("## Totals (all files)")
    lines.append("")
    lines.append("| table | inserted | updated | skipped |")
    lines.append("|---|---:|---:|---:|")
    for table_name in sorted(totals.keys()):
        lines.append(_row(table_name, totals[table_name]))
    lines.append("")
    return "\n".join(lines)


# PUBLIC_INTERFACE
def run() -> int:
    """CLI entrypoint for Step 2.2–2.5 ingestion and consolidated reporting."""
    parser = argparse.ArgumentParser(description="CloudUnify Pro - Step 2.2→2.5 ingestion (Neon)")
    parser.add_argument("--database-url", default="", help="Neon DATABASE_URL (postgresql://...)")
    parser.add_argument("--org-name", default="CloudUnify Demo")
    parser.add_argument("--org-slug", default="cloudunify-demo")

    parser.add_argument("--aws-file", default="/home/kavia/workspace/code-generation/attachments/20251220_123809_mock_aws_resources.xlsx")
    parser.add_argument("--azure-file", default="/home/kavia/workspace/code-generation/attachments/20251220_123811_mock_azure_resources.xlsx")
    parser.add_argument("--gcp-file", default="/home/kavia/workspace/code-generation/attachments/20251220_123813_mock_gcp_resources.xlsx")
    parser.add_argument("--recommendations-file", default="/home/kavia/workspace/code-generation/attachments/20251220_123814_mock_recommendations.xlsx")

    parser.add_argument("--account-name-aws", default="AWS Main")
    parser.add_argument("--account-name-azure", default="Azure Main")
    parser.add_argument("--account-name-gcp", default="GCP Main")

    parser.add_argument("--report-md", default="/home/kavia/workspace/code-generation/logs/import_report_final_20251220.md")
    parser.add_argument("--report-json", default="/home/kavia/workspace/code-generation/logs/import_report_final_20251220.json")

    args = parser.parse_args()

    raw_dsn = args.database_url or os.getenv("DATABASE_URL", "")
    if not raw_dsn:
        raise SystemExit("ERROR: provide --database-url or set DATABASE_URL")

    dsn = _ensure_ssl_and_channel_binding(raw_dsn)

    started_at = _now_utc()

    resource_paths = [
        ("aws", Path(args.aws_file)),
        ("azure", Path(args.azure_file)),
        ("gcp", Path(args.gcp_file)),
    ]
    rec_path = Path(args.recommendations_file)

    for prov, p in resource_paths:
        if not p.exists():
            raise SystemExit(f"ERROR: missing {prov} resources file: {p}")
    if not rec_path.exists():
        raise SystemExit(f"ERROR: missing recommendations file: {rec_path}")

    # Ensure output directories exist
    Path(args.report_md).expanduser().parent.mkdir(parents=True, exist_ok=True)
    if args.report_json:
        Path(args.report_json).expanduser().parent.mkdir(parents=True, exist_ok=True)

    lookups: Dict[str, TableCounts] = {"organizations": TableCounts(), "cloud_accounts": TableCounts()}
    file_reports: List[FileReport] = []

    with psycopg.connect(dsn, autocommit=False) as conn:
        has_rcd = _table_exists(conn, "resource_costs_daily")

        # Lookup upserts
        org_id, org_inserted = _upsert_org(conn, args.org_name, args.org_slug)
        if org_inserted:
            lookups["organizations"].inserted += 1
        else:
            lookups["organizations"].updated += 1

        cloud_account_ids: Dict[str, str] = {}
        for prov, account_name in (
            ("aws", args.account_name_aws),
            ("azure", args.account_name_azure),
            ("gcp", args.account_name_gcp),
        ):
            acc_id, acc_inserted = _upsert_cloud_account(
                conn,
                organization_id=org_id,
                provider=prov,
                account_external_id=f"mock-{prov}",
                account_name=account_name,
            )
            cloud_account_ids[prov] = acc_id
            if acc_inserted:
                lookups["cloud_accounts"].inserted += 1
            else:
                lookups["cloud_accounts"].updated += 1

        conn.commit()

        # Resources
        for prov, pth in resource_paths:
            fr = _ingest_resources_xlsx(
                conn,
                organization_id=org_id,
                provider=prov,
                cloud_account_id=cloud_account_ids[prov],
                path=pth,
                has_resource_costs_daily=has_rcd,
            )
            file_reports.append(fr)

        # Recommendations
        fr_rec = _ingest_recommendations_xlsx(
            conn,
            organization_id=org_id,
            cloud_account_ids=cloud_account_ids,
            path=rec_path,
        )
        file_reports.append(fr_rec)

    finished_at = _now_utc()
    totals = _sum_table_counts(file_reports)

    org_payload = {
        "id": org_id,
        "name": args.org_name,
        "slug": args.org_slug,
        "cloud_accounts": {
            "aws": {"id": cloud_account_ids["aws"], "name": args.account_name_aws, "account_external_id": "mock-aws"},
            "azure": {"id": cloud_account_ids["azure"], "name": args.account_name_azure, "account_external_id": "mock-azure"},
            "gcp": {"id": cloud_account_ids["gcp"], "name": args.account_name_gcp, "account_external_id": "mock-gcp"},
        },
    }

    md = _render_markdown_report(
        started_at=started_at,
        finished_at=finished_at,
        database_url=dsn,
        org=org_payload,
        lookups=lookups,
        file_reports=file_reports,
        totals=totals,
    )
    Path(args.report_md).write_text(md, encoding="utf-8")

    # JSON report (if requested)
    if args.report_json:
        json_report = {
            "started_at": started_at.isoformat(),
            "finished_at": finished_at.isoformat(),
            "database_url": dsn,
            "organization": org_payload,
            "lookups": {k: asdict(v) for k, v in lookups.items()},
            "files": [
                {
                    "file": fr.file,
                    "provider": fr.provider,
                    "tables": {t: asdict(c) for t, c in fr.tables.items()},
                }
                for fr in file_reports
            ],
            "totals": {t: asdict(c) for t, c in totals.items()},
        }
        Path(args.report_json).write_text(json.dumps(json_report, indent=2), encoding="utf-8")

    # Console summary
    print(f"Wrote report: {args.report_md}")
    if args.report_json:
        print(f"Wrote report: {args.report_json}")
    print("Totals:", {k: asdict(v) for k, v in totals.items()})

    return 0


if __name__ == "__main__":
    raise SystemExit(run())
